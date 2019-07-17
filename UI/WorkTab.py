#每个tab包含两个view用于显示源文件与目标文件
import os
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import openpyxl
import re
import xml.etree.cElementTree as XETree
import win32clipboard
import win32con
import numpy as np

class WorkTab(QTabWidget):

    def __init__(self):
        super().__init__()
        self.__init()

    def __init(self):
        self.setTabsClosable(True)
        self.setMovable(True)
        self.tabCloseRequested.connect(lambda index: self.removeTab(index)
                                       )
        #设置最大宽度
        self.setStyleSheet("QTabBar::tab{max-width:150px}")
        #self.tabBar().tabBarClicked.connect(lambda index: print(("this is %d" % index))
        # str = "QTabBar::tab{background-color:rbg(255,255,255,0);}" + \
        #       "QTabBar::tab:selected{color:red;background-color:rbg(255,200,255);} "
        self.tabBar().setContextMenuPolicy(Qt.CustomContextMenu)
        self.tabBar().customContextMenuRequested.connect(self.__showMenu)
        self.contextMenu = QMenu(self)
        self.contextMenu.setStyleSheet("background-color:rbg(0,0,255);"
                                       "selection-background-color: rgb(85, 170, 255);")

        self.action_CloseAll = QAction("关闭所有标签")
        self.action_CloseLeft = QAction("关闭左侧标签")
        self.action_CloseRight = QAction("关闭右侧标签")
        self.action_CloseAllButThis = QAction("关闭除当前标签以外的所有标签")

        self.contextMenu.addAction(self.action_CloseAll)
        self.contextMenu.addAction(self.action_CloseLeft)
        self.contextMenu.addAction(self.action_CloseRight)
        self.contextMenu.addAction(self.action_CloseAllButThis)

        self.action_CloseAll.triggered.connect(self.clear)
        self.action_CloseLeft.triggered.connect(self.__removeLeftTab)
        self.action_CloseRight.triggered.connect(self.__removeRightTab)
        self.action_CloseAllButThis.triggered.connect(self.__removeAllButThis)


    def __showMenu(self):
        self.contextMenu.exec_(QCursor.pos())

    def __removeLeftTab(self):
        index = self.currentIndex()
        for i in range(0, index):
            result = self.removeTab(0)

    def __removeRightTab(self):
        index = self.currentIndex()
        for i in range(index + 1, self.count()):
            self.removeTab(index + 1)
    def __removeAllButThis(self):
        self.__removeLeftTab();
        self.__removeRightTab()

    def removeTab(self, index: int) -> None:
        tab = self.widget(index)
        if tab is not None:
            str = ""
            saved = True
            if tab.isChanged:
                str += "\n" + self.tabText(index)
                saved = False
            if not saved:
                reply = QMessageBox.question(self, '确认', \
                                             '以下任务结果还未保存，确认退出？' + str, \
                                             QMessageBox.Yes | QMessageBox.No, \
                                             QMessageBox.No)

                if reply == QMessageBox.Yes:
                    super().removeTab(index)
                    tab.setParent(None)
                    tab = None
                    return True
            else:
                super().removeTab(index)
                tab.setParent(None)
                tab = None
    def copy(self):
        self.currentWidget().copy()

    def paste(self):
        self.currentWidget().paste()

    def clear(self) -> None:
        self.__removeAllButThis()
        self.removeTab(self.currentIndex())

    def setTabIconByStatus(self, widget, status: bool):
        if not status:
            self.setTabIcon(self.indexOf(widget), QIcon("Resource/icon/Icon_tag.ico"))
        else:
            self.setTabIcon(self.indexOf(widget), QIcon("Resource/icon/Icon_UnSave.png"))

    def changeEvent(self, event: QEvent) -> None:
        return

    def addTab(self, widget: QWidget, icon: QIcon,  title: str) -> int:
        super().addTab(widget,icon, title)
        widget.signal_changed.connect(self.setTabIconByStatus)




class Tab(QWidget):
    essentialColor = QColor(255, 230, 153)
    isChanged = None
    logGenerated = pyqtSignal(str)
    signal_changed = pyqtSignal(QWidget, bool)


    def __init__(self):
        super().__init__()
        self.initUI()

    def initProperty(self, taskName, filepath, outputfolder, xmlName, xmlFIlePath, templateFileName, templateFilePath, isDirectOutput):
        self.taskName = taskName
        self.outputfolder = outputfolder
        self.xmlName = xmlName
        self.xmlFIlePath = xmlFIlePath
        self.templateFileName = templateFileName
        self.templateFilePath = templateFilePath
        self.isDirectOutput = isDirectOutput
        self.filepath = filepath
        self.destFilepath = ""
        self.titleBar.setText("任务名: {0} 方案：{1} 模板文件： {2}   输出至：{3}".format(self.taskName, self.xmlName,
                              self.templateFileName,
                              self.outputfolder))

    def initUI(self):
        # 记录哪些单元格被标记了颜色
        self.souce_markItem = []
        self.dest_markItem = []

        self.setStyleSheet("background:white;")
        self.sourceWidget = QWidget()
        self.destWidget = QWidget()
        self.sourceSheet = Sheet()
        self.destSheet = Sheet()
        self.__map = dict()

        self.gridLayout = QGridLayout()
        self.setLayout(self.gridLayout)

        self.titleBar = QLabel()

        self.gridLayout.addWidget(self.titleBar, 0, 0)
        self.horizontalSplitter = QSplitter(Qt.Horizontal, self)
        self.horizontalSplitter.addWidget(self.sourceWidget)
        self.horizontalSplitter.addWidget(self.destWidget)
        self.horizontalSplitter.setStyleSheet("height:100%")

        self.gridLayout .addWidget(self.horizontalSplitter, 1, 0)
        # table = Widget()
        # self.gridLayout.addWidget(table, 0, 1)

        #self.vbox.setDirection(0)

        self.sbox = QVBoxLayout()
        self.sbox.addWidget(self.sourceSheet)
        self.sourceWidget.setLayout(self.sbox)

        self.dbox = QVBoxLayout()
        self.dbox.addWidget(self.destSheet)
        self.destWidget.setLayout(self.dbox)

        #添加右键粘贴菜单
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self.__showMenu)
        self.contextMenu = QMenu(self)
        self.contextMenu.setStyleSheet("background-color:rbg(0,0,255);"
                                       "selection-background-color: rgb(85, 170, 255);")

        self.action_PasteNotFormat = QAction("粘贴（保留原格式）")
        self.action_PasteAndTranspositionNotFormat = QAction("转置粘贴（保留原格式）")
        self.action_PasteAndTranspositionAndFormat = QAction("转置粘贴（匹配目标格式）")

        self.contextMenu.addAction(self.action_PasteNotFormat)
        self.contextMenu.addAction(self.action_PasteAndTranspositionNotFormat)
        self.contextMenu.addAction(self.action_PasteAndTranspositionAndFormat)

        self.action_PasteNotFormat.triggered.connect(lambda: self.paste(False, False))
        self.action_PasteAndTranspositionNotFormat.triggered.connect(lambda: self.paste(False, True))
        self.action_PasteAndTranspositionAndFormat.triggered.connect(lambda: self.paste(True, True))

        self.sourceSheet.mouseReleased.connect(self.focusLeftToRight)
        self.destSheet.mouseReleased.connect(self.focusRihtToLeft)

        self.sourceSheet.setStyleSheet("QTableWidget::item::selected{border:3px solid rgb(255,0,0);"
                                       "background-color:rgba(255,0,0,130)};")
        self.destSheet.setStyleSheet("QTableWidget::item::selected{border:3px solid rgb(0,0,255);"
                                     "background-color:rgba(0,0,255,130)};")
        self.titleBar.setStyleSheet("margin-top:3px;margin-bottom:3px;max-height:30px")
        self.titleBar.setText("新任务")
        # self.sourceWidget.setStyleSheet("background-color: rgb(85, 170, 255);")
        # self.destWidget.setStyleSheet("background-color: rgb(255, 135, 137);")
    def run(self):
        outputFolder = self.outputfolder
        xmlfilepath = self.xmlFIlePath
        templatefile = self.templateFilePath
        inputFile = self.filepath
        outputFile = os.path.join(outputFolder, os.path.basename(inputFile))
        if not os.path.isfile(self.filepath):
            QMessageBox.warning(self, '警告', '请先点击新建任务', QMessageBox.Yes)
            return
        self.logGenerated.emit("开始转换文件：{0} 方案：{1} 模板文件：{2} 输出至：{3}".format(os.path.basename(inputFile),
                                                                           self.xmlName,
                                                                           self.templateFileName,
                                                                           outputFolder))

        self.beginTranslate(inputFile, xmlfilepath, templatefile)
        self.logGenerated.emit("转换完成")
        self.signal_changed.emit(self, True)
        if self.isDirectOutput:
            self.save()


    def copy(self):
        widget = self.focusWidget()
        if type(widget) == Sheet:
            indexes = widget.selectedIndexes()
            #复制到剪切板
            #确定选择的行列
            rowset = set()
            for index in indexes:
                rowset.add(index.row())
            rowset = list(rowset)
            #排序
            for i in range(len(rowset) - 1):
                for j in range(len(rowset) - i -1):
                    if rowset[j] > rowset[j + 1]:
                        temp = rowset[j + 1]
                        rowset[j + 1] = rowset[j]
                        rowset[j] = temp


            colset = set()
            for index in indexes:
                colset.add(index.column())
            colset = list(colset)

            for i in range(len(colset) - 1):
                for j in range(len(colset) - i -1):
                    if colset[j] > colset[j + 1]:
                        temp = colset[j + 1]
                        colset[j + 1] = colset[j]
                        colset[j] = temp

            #拼接字符
            clip_str = ""
            for i in range(len(rowset)):
                if i != 0:
                    clip_str += "\r\n"
                for j in range(len(colset)):
                    if j != 0:
                        clip_str += "\t"
                    # 添加单元格文本
                    item = widget.item(rowset[i], colset[j])
                    if item is not None:
                        clip_str += item.text()
                    else:
                        clip_str += ""

            # #去掉一开始多加的\t
            # clip_str = clip_str[1:]
             #对于原表格，额外追加复制的单元格
            clip_str += "/s/s/s"

            clip_coordinate = ""
            for i in range(len(rowset)):
                if i != 0:
                    clip_coordinate += "\r\n"
                for j in range(len(colset)):
                    if j != 0:
                        clip_coordinate += "\t"
                    clip_coordinate += "{0},{1}".format(rowset[i], colset[j])

            clip_str += clip_coordinate
            #发送到剪切板
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32con.CF_UNICODETEXT, clip_str)
            win32clipboard.CloseClipboard()
        return
            
    def paste(self, format = True, transposition = False):
        widget = self.focusWidget()
        #判断是否是表格
        if  type(widget) == Sheet:
            win32clipboard.OpenClipboard()
            text = win32clipboard.GetClipboardData(win32con.CF_UNICODETEXT)
            win32clipboard.CloseClipboard()


            pattern = text.split("/s/s/s")
            rows = pattern[0].split("\r\n")
            cell = []
            # 解析数据
            for i in range(len(rows)):
                cell.append(rows[i].split("\t"))
            # 开始填充
            # item = widget.selectedItems()[0]
            startrow = widget.selectedIndexes()[0].row()
            startcol = widget.selectedIndexes()[0].column()
            #是否转置
            if transposition:
                cell = np.array(cell).T
            if self.sourceSheet == widget:
                widget.pasteData(startrow, startcol, cell, False)
            else:
                widget.pasteData(startrow, startcol, cell, format)

            #修改源和目标的map
            if self.destSheet == self.focusWidget() and len(pattern) > 1:
                rows_item = pattern[1].split("\r\n")
                items = []
                # 解析数据
                for i in range(len(rows_item)):
                    items.append(rows_item[i].split("\t"))
                if transposition:
                    items = zip(*items)

                self.addMapData(startrow, startcol, items)
        return
    def __showMenu(self):
        self.contextMenu.exec_(QCursor.pos())

    def addMapData(self, startrow, startcol, data):
        i = 0
        for row in data:
            for j in range(len(row)):
                key = "{0},{1}".format(startrow + i, startcol + j)
                self.addMapValue(key, row[j])
            i += 1
        self.refreshSourceSheet()

    def addMapValue(self, key, value):
        if value in self.__map.values():
            self.__map.pop(list(self.__map.keys())[list(self.__map.values()).index(value)])
        self.__map[key] = value

    def focusRihtToLeft(self):
        self.sourceSheet.clearSelection()
        indexes = self.destSheet.selectedIndexes()
        for index in indexes:
            key = "{0},{1}".format(index.row(), index.column())
            if key in self.__map:
                sItem = self.__map[key]
                if sItem != "" and sItem != "NA":
                    item = self.sourceSheet.item(int(sItem.split(",")[0]), int(sItem.split(",")[1]))
                    if item is None:
                        item = TableItem("")
                        self.sourceSheet.setItem(int(sItem.split(",")[0]), int(sItem.split(",")[1]), item)
                    self.sourceSheet.scrollToItem(item, QAbstractItemView.PositionAtCenter)
                    self.sourceSheet.setRangeSelected(
                        QTableWidgetSelectionRange(item.row(), item.column(), item.row(), item.column()), True)
    def focusLeftToRight(self):
        self.destSheet.clearSelection()
        indexes = self.sourceSheet.selectedIndexes()
        for i in range(len(indexes)):
            value = "{0},{1}".format(indexes[i].row(),indexes[i].column())
            if value in self.__map.values():
                key = list (self.__map.keys()) [list (self.__map.values()).index (value)]
                item = self.destSheet.item(int(key.split(",")[0]), int(key.split(",")[1]))
                if item is None:
                    item = TableItem("")
                    self.destSheet.setItem(int(key.split(",")[0]), int(key.split(",")[1]), item)
                self.destSheet.scrollToItem(item, QAbstractItemView.PositionAtCenter)
                self.destSheet.setRangeSelected(
                    QTableWidgetSelectionRange(item.row(), item.column(), item.row(), item.column()), True)
        return


    def fillLeft(self, sheetIndex):
        self.sourceSheet.fillSheetByExcelSheetIndex(self.filepath, sheetIndex)

    def fillRight(self, filepath, sheetIndex):
        self.destSheet.fillSheetByExcelSheetIndex(filepath, sheetIndex)

    def beginTranslate(self, filePath, xmlFilePath, templateFilePath):
        if not self.destSheet.hasLoad():
            self.fillRight(templateFilePath, 0)
        try:
            cfgItems = XETree.parse(xmlFilePath).getroot()
            for i in range(len(cfgItems)):
                cfgItem = cfgItems[i]
                istable = cfgItem.attrib['istable'].lower()
                if istable == 'true':
                    self.__extractForTable(cfgItem)
                else:
                    self.__extractForKeyValue(cfgItem)
            self.refreshSourceSheet()
        except Exception as e:
            QMessageBox.critical(self, "错误", "xml文件解析错误" + str(e), QMessageBox.Ok)


    def __extractForKeyValue(self, cfgItem):
        sNode = cfgItem.find('source')
        dNode = cfgItem.find('dest')
        dCols = dNode.attrib['cols']
        anchors = sNode.attrib['anchor'].strip().split(';')
        length = len(anchors)
        beginCol = 0
        beginRow = 0
        sCols = sNode.attrib['cols']
        isFind = 0;
        datatype = dNode.attrib['datatype']
        tempcell = None
        for i in range(length):
            tempcell = self.__findStr(self.sourceSheet, anchors[i], beginRow, beginCol)
            if tempcell is not None:
                beginRow = tempcell.row()
                beginCol = tempcell.column() + 1
                if i == length - 1:
                    isFind = 1
        if isFind == 1:
            sourceItem = self.sourceSheet.item(tempcell.row(), letterToint(sCols))
            self.addMapValue(parseChar(dCols), str(sourceItem.row()) + "," + str(sourceItem.column()))
            item = self.destSheet.item(int(re.sub(r"\D", "", dCols)) - 1, letterToint(re.sub(r"[^A-Z]", "", dCols)))
            item.dataType = datatype
            item.setFormatValue(sourceItem.text())
            #item.setBackground(self.essentialColor)
        if isFind ==0:
            item = self.destSheet.item(int(re.sub(r"\D", "", dCols)) - 1, letterToint(re.sub(r"[^A-Z]", "", dCols)))
            item.dataType = datatype
            item.setFormatValue("")
            item.setBackground(self.essentialColor)

    def __extractForTable(self, cfgItem):
        endrow = 0
        beginrow = 0
        error_str = ""
        tempcell = self.__findStr(self.sourceSheet, cfgItem[0].attrib["anchor"], 0, 0)
        if tempcell is None:
            return
        self.sourceSheet.rowsMark.append(tempcell.row() - 1)
        destBeginRow = int(cfgItem[1].attrib["beginrow"])
        sList = cfgItem[0].attrib["cols"].split(',')
        dList = cfgItem[1].attrib["cols"].split(',')
        datatype = cfgItem[1].attrib["datatype"].split(',')
        if len(sList) != len(dList):
            error_str ="{0} :原列数{1}与目标列数{2}不一致".format( cfgItem.attrib["desc"], len(sList), len(dList))
        elif len(sList) != len(datatype):
            error_str ="{0} :原列数{1}与目标类型数{2}不一致".format( cfgItem.attrib["desc"], len(sList), len(datatype))
        elif len(dList) != len(datatype):
            error_str ="{0} :目标列数{1}与目标类型数{2}不一致".format( cfgItem.attrib["desc"], len(dList), len(datatype))
        if error_str != "":
            QMessageBox.warning(self, "错误", error_str, QMessageBox.Ok)
            return
        for col in range(len(dList)):
            for row in range(int(cfgItem[1].attrib["limited"])):
                item = self.destSheet.item(destBeginRow + row -1, letterToint(dList[col]))
                if item is not None:
                    #item.setBackground(self.essentialColor)
                    item.dataType = datatype[col]
        #如果差找不到，首行设置为NA
        if tempcell is None:
            for col in range(len(dList)):
                item = self.destSheet.item(destBeginRow - 1, letterToint(dList[col]))
                #item.dataType = datatype[col]
                item.setText("NA")
            return
        # 判断在源文件中的起始行
        beginrow = tempcell.row() + int(cfgItem[0].attrib["skiprows"]) + 1

        # 判断结束行
        # 如果范围已经确定，直接确定结束行
        if cfgItem[0].attrib["range"] != "":
            endrow = beginrow + int(cfgItem[0].attrib["range"]) - 1
        # 范围不确定，根据下一行字符确定结束行
        elif cfgItem[0].attrib["anchorend"] != "":
            endcell = self.__findStr(self.sourceSheet, cfgItem[0].attrib["anchorend"], beginrow, 0)
            if endcell is not None:
                endrow = endcell.row()
            else:
                return
        else:
            # 找不到字符，则直到最后一个不为空行的为止
            limited = 1
            while self.sourceSheet.item(beginrow + limited, 0).text() != "" and self.sourceSheet.item(beginrow + limited, 0).text() is not None:
                limited += 1
            endrow = beginrow + limited - 1

        #print(cfgItem[0].attrib["anchor"])
        # 粘贴数据
        for row in range(beginrow, endrow + 1):
            for col in range(0, len(sList)):
                sourceItem = self.sourceSheet.item(row, letterToint(sList[col]))
                self.addMapValue(parseChar(dList[col] + str(destBeginRow + row - beginrow)), str(sourceItem.row()) + "," + str(sourceItem.column()))
                item = self.destSheet.item(destBeginRow + row - beginrow - 1, letterToint(dList[col]))
                #item.dataType = datatype[col]
                item.setFormatValue(sourceItem.text())

    def __findStr(self, sheet, key, startRow, startCol):
        for col in range(startCol, sheet.columnCount()):
            for row in range(startRow, sheet.rowCount()):
                cell = sheet.item(row, col)
                if cell is None:
                    return None
                if cell.text() is not None and cell.text() != "":
                    if key in str(cell.text()).replace(' ', ''):
                        return cell

    def refreshSourceSheet(self):
        for cordinnate in self.souce_markItem:
            self.sourceSheet.item(int(cordinnate.split(",")[0]), int(cordinnate.split(",")[1])).setBackground(QColor(255, 255, 255))
        for cordinnate in self.dest_markItem:
            self.destSheet.item(int(cordinnate.split(",")[0]), int(cordinnate.split(",")[1])).checkData(QColor(255, 255, 255))
        for (k, v) in self.__map.items():
            k_x = int(k.split(",")[0])
            k_y = int(k.split(",")[1])
            v_x = int(v.split(",")[0])
            v_y = int(v.split(",")[1])
            vItem = self.sourceSheet.item(v_x, v_y)
            kItem = self.destSheet.item(k_x, k_y)
            if vItem is None:
                vItem = TableItem("")
                self.sourceSheet.setItem(v_x, v_y, vItem)
            if kItem is None:
                kItem = TableItem("")
                self.destSheet.setItem(k_x, k_y, kItem)
            vItem.setBackground(QColor(85, 170, 255))
            kItem.checkData(self.essentialColor)
            self.souce_markItem.append(v)
            self.dest_markItem.append(k)


    def save(self):
        filepath = self.destFilepath
        if filepath == "":
            if self.isDirectOutput:
                outputdir = self.outputfolder
                if not os.path.isdir(outputdir):
                    reply = QMessageBox.warning(self, "警告", "请选择正确的输出路径", QMessageBox.Ok)
                    if reply == QMessageBox.Ok:
                        return None
                else:
                    filepath = os.path.join(outputdir, os.path.basename(self.filepath))
            else:
                fname = QFileDialog.getSaveFileName(self, '保存文件', self.filepath, filter="*.xlsx",)
                if (fname[0]):
                    filepath = fname[0]
                else:
                    return None

        templateFile = self.templateFilePath
        if os.path.isfile(templateFile):
            open(filepath, "wb").write(open(templateFile, "rb").read())
            self.__save(filepath, 0)
            self.logGenerated.emit("已保存至： " + filepath)
            self.setTabStatus(False)
        else:
            QMessageBox.warning(self, "警告", "请选择正确的模板文件", QMessageBox.Ok)


    def __save(self, filepath, sheetIndex):
        if self.isChanged is not None:
            self.destFilepath = filepath
            workbok = openpyxl.load_workbook(filepath)
            sheetNames = workbok.sheetnames
            sheet = workbok[sheetNames[sheetIndex]]
            for col in range(0, self.destSheet.columnCount()):
                for row in range(0, self.destSheet.rowCount()):
                    cell = self.destSheet.item(row, col)
                    if cell is None:
                        continue
                    if cell.text() is not None and cell.text() != "":
                        try:
                            value = cell.text()
                            if cell.dataType == "F":
                                value = float(value)
                            sheet[intToletter(col) + str(row + 1)].value = value
                        except Exception as e:
                            print(intToletter(col) + str(row))
                            print(e)
            workbok.save(filepath)

    def keyPressEvent(self, event: QKeyEvent):
        super().keyPressEvent(event)
        if event.key() == Qt.Key_Delete:
            widget = self.focusWidget()
            if type(widget) == Sheet:
                indexes = widget.selectedIndexes()
                for index in indexes:
                    item = widget.itemFromIndex(index)
                    if item is not None:
                        item.setText("")
                        item.setBackground(QColor(255, 255, 255))
                        coordinnate = "{0},{1}".format(item.row(), item.column())
                    if widget == self.destSheet:
                         if coordinnate in self.__map:
                             self.__map.pop(coordinnate)
                    if widget == self.sourceSheet:
                         if coordinnate in self.__map.values():
                             self.__map.pop(list(self.__map.keys())[list(self.__map.values()).index(coordinnate)])

    def setTabStatus(self, status):
        self.isChanged = status
        self.signal_changed.emit(self, status)


class Sheet(QTableWidget):
    rowsMark = [int]
    mouseReleased = pyqtSignal()
    def __init__(self):
        super().__init__()
        self.__hasload = False
        self.rowsMark.append(0)
        self.setColumnCount(10)
        self.setRowCount(30)
        for i in range(0, 10):
            item = TableItem(value=intToletter(i))
            self.setHorizontalHeaderItem(i, item)
        for i in range(0, 30):
            item = TableItem(value=str(i+1))
            self.setVerticalHeaderItem(i, item)
        self.horizontalScrollBar().setStyleSheet("height:25px;background-color: rgb(222, 222, 222);")
        self.verticalScrollBar().setStyleSheet("width:25px;background-color: rgb(222, 222, 222);")
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)
    def mouseReleaseEvent(self, event: QMouseEvent) -> None:
        super().mouseReleaseEvent(event)
        self.mouseReleased.emit()
        event.accept()
    
    def fillSheetByExcelSheetIndex(self, filepath, sheetIndex):
        try:
            workbok = openpyxl.load_workbook(filepath)
            sheetNames = workbok.sheetnames
            self.__fileSheetByWorkBook(workbok[sheetNames[sheetIndex]])
        except Exception as e:
            QMessageBox.critical(self, "错误", "解析文件发生错误: " + str(e), QMessageBox.Ok)

    def fillSheetByExcelSheetName(self, filepath, sheetName):
        try:
            workbok = openpyxl.load_workbook(filepath)
            self.__fileSheetByWorkBook(workbok[sheetName])
        except Exception as e:
            QMessageBox.critical(self, "错误", "解析文件发生错误：" + str(e), QMessageBox.Ok)

    #设置item位置超过table大小时自动更新大小
    def setItem(self, row: int, column: int, item: QTableWidgetItem) -> None:
        #例如当前row = 9，实际rowcount = 9，rowcount应该改为10， 并给新hearder设置值
        current_rowCount = self.rowCount()
        current_columnCount = self.columnCount()
        if row > current_rowCount  - 1:
            self.setRowCount(row + 1)
            for i in range(current_rowCount, row + 1):
                self.setVerticalHeaderItem(i, TableItem(str(i+1)))
        if column > current_columnCount  - 1:
            self.setColumnCount(column + 1)
            for i in range(current_columnCount, column + 1):
                self.setHorizontalHeaderItem(i, TableItem(intToletter(i)))
        super().setItem(row, column, item)



    def __fileSheetByWorkBook(self, sheet):
        row_count = sheet.max_row
        col_count = sheet.max_column
        if self.rowCount() < row_count:
            self.setRowCount(row_count)
        if self.columnCount() < sheet.max_column:
            self.setColumnCount(col_count)
        for i in range(0, col_count):
            if self.horizontalHeaderItem(i) is None:
                item = TableItem(value=intToletter(i))
                self.setHorizontalHeaderItem(i, item)
        for i in range(0, row_count):
            if self.verticalHeaderItem(i) is None:
                item = TableItem(value=str(i + 1))
                self.setVerticalHeaderItem(i, item)

        #开始初始化表格
        for row in sheet.iter_rows(min_row=0):
            for cell in row:
                item = TableItem("")
                if type(cell.column) == str:
                    self.setItem(cell.row - 1, letterToint(cell.column), item)
                elif type(cell.column) == int:
                    self.setItem(cell.row - 1, cell.column - 1, item)
                # color = cell.fill.bgColor.rgb
                # if color is not None and color != "00000000" and type(color) != openpyxl.styles.colors.RGB:
                #     color = HexToRgb(color)
                #     item.setBackground(QColor(color["r"], color["g"], color["b"]))
                if cell.value is not None and cell.value != "":
                    item.setText(str(cell.value))

        #合并单元格
        merged_cell = str(sheet.merged_cells).split(" ")
        for cell in merged_cell:
            if cell != "":
                startPoint = cell.split(":")[0]
                endPoint = cell.split(":")[1]
                startRow = int(re.sub(r"\D", "", startPoint)) - 1
                startCol = letterToint(re.sub(r"[^A-Z]", "", startPoint))
                endRow = int(re.sub(r"\D", "", endPoint)) - 1
                endCol = letterToint(re.sub(r"[^A-Z]", "", endPoint))
                self.setSpan(startRow, startCol, endRow - startRow + 1, endCol - startCol + 1)
        self.__hasload = True

    def pasteData(self, startrow, startcol, data, format=True):
        i = 0
        for row in data:
            for j in range(len(row)):
                item = self.item(startrow + i, startcol + j)
                if item is None:
                    item = TableItem("")
                    self.setItem(startrow + i, startcol + j, item)
                if format:
                    item.setFormatValue(row[j])
                else:
                    item.setText(row[j])
            i += 1

    def hasLoad(self):
        return self.__hasload

    def setColumnCount(self, columns: int) -> None:
        current_column = self.columnCount()
        super().setColumnCount(columns)
        if columns > current_column:
            for i in range(current_column, columns):
                self.setColumnWidth(i, 110)




class TableItem(QTableWidgetItem):
    dataType = "S"
    IsSaved = True
    errorColor = QBrush(QColor(255, 114, 116))
    noramlColor = QBrush(QColor(255, 255, 255))
    essentialColor = QColor(255, 230, 153)
    def __int__(self):
        super().__init__()
        self.__init()
    def __init__(self, value):
        super().__init__()
        self.__init()
        self.setText(value)
    def __init(self):
        self.setForeground(QColor(0, 0, 0))
    # def __init__(self, dataType, value):
    #     super().__init__()
    #     self.setFormatValue(value)
    def setFormatValue(self,value):
        value = str(value)
        try:
            value = self.updateValue(value)
            super().setText(value)
            self.checkData(self.essentialColor)
        except Exception as e:
            print(e)

    def updateValue(self,value):  # 根据类型进行字符串处理，D日期，P百分数，S字符串不做处理，F或不填按数字处理
        value = str(value)
        try:
            if value is None:
                return ''
            value = str(value)
            value = value.replace(' ', '')
            if self.dataType == 'D':
                # dateFormat = '[^\d/\d/\d]'
                # '2016年3月1日至2016年3月31日'
                value = re.sub('.*至', '', value)
                value = re.sub('\D$', '', value)
                value = re.sub(r'\D', r'/', value)
            elif self.dataType == 'P':
                if type(value) == float:
                    value = str(value * 100) + '%'
                else:
                    Persentage = '[^\d%|\d.\d%]'
                    value = re.sub(Persentage, '', value)
            elif self.dataType == 'S':
                return value
            else:
                floatFormat = '[^\d|\d.\d]'
                value = re.sub(floatFormat, '', value)
                lenth = len(value.split('.')) - 2
                value = value.replace('.', '', lenth)
                if value == '' or value == '.':
                    value = "0"
            return value
        except Exception as e:
            print(e)
            print("ErrorValue: " + value + "Type: " + self.ataType)

    def checkData(self, normalColor):  # 值校验，是否符合对应类型
        value = str(self.text())
        date = '^\d{4}\D\d{1,2}\D\d{1,2}\D?$'
        num = '^[-]?\d*[.]?\d*$'
        persentige = '^[-]?\d*[.]?\d*[%]$'
        if value is None or (value == '' and self.dataType != "S"):
            self.setBackground(self.errorColor)
            return False
        flag = 0
        if self.dataType == 'D':
            flag = re.search(date, value)
        elif self.dataType == 'P':
            flag = re.search(persentige, value)
        elif self.dataType == 'S':
            self.setBackground(normalColor)
            return True
        else:
            flag = re.search(num, value)
        if str(flag) == 'None':
            self.setBackground(self.errorColor)
            return False
        else:
            self.setBackground(normalColor)
            return True
        return flag


    def getDataType(self):
        return self.dataType
    def setDataType(self, dataType):
        self.dataType = dataType



#列字母转数字
def letterToint(s):
    letterdict = {}
    for i in range(26):
        letterdict[chr(ord('A') + i)] = i + 1
    output = 0
    for i in range(len(s)):
        output = output * 26 + letterdict[s[i]]
    return output - 1
#数字转列字母
def intToletter(i):
    if type(i) is not int:
        return i
    str = ''
    i += 1
    while (not (i // 26 == 0 and i % 26 == 0)):

        temp = 25

        if (i % 26 == 0):
            str += chr(temp + 65)
        else:
            str += chr(i % 26 - 1 + 65)

        i //= 26
        # print(str)
    # 倒序输出拼写的字符串
    return str[::-1]

def parseChar(string):
    letter = re.sub(r"[^A-Z]", "", string)
    num = int(re.sub(r"\D", "", string)) -1
    return "{0},{1}".format(str(num),letterToint(letter))
#十六进制颜色编码转rgb
def HexToRgb(tmp):
    rgb = dict()
    opt = re.findall(r'(.{2})',tmp)
    rgb["r"] = 255 - int(opt[0], 16)
    rgb["g"] = 255 - int(opt[1], 16)
    rgb["b"] = 255 - int(opt[2], 16)
    return rgb
